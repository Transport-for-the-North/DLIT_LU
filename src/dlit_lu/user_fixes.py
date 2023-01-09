# standard imports
import pathlib
import os
import logging
from typing import Optional
# third party imports
import pandas as pd
import openpyxl

# local imports
from dlit_lu import global_classes, utilities, parser, analyse, inputs, data_repair

LOG = logging.getLogger(__name__)


def user_input_file_builder(path: pathlib.Path, data: global_classes.DLogData) -> None:
    """builds file for user to edit 

    file can then be read and edits integrated into data

    Parameters
    ----------
    path : pathlib.Path
        location to save file
    data : global_classes.DLogData
        data to save
    """
    data = utilities.to_dict(data)
    utilities.write_to_excel(path, data)


def infill_user_inputs(
    data: dict[str, pd.DataFrame],
    modified_path: pathlib.Path,
    uneditable_columns: Optional[dict[str, list[str]]] = None,
) -> dict[str, pd.DataFrame]:
    """read in user edited file and integrate edits into data

    will fail if user edits column names or indices
    should work if modified path contains a subset of data as
    long as indices are consistent 

    Parameters
    ----------
    data : dict[str, pd.DataFrame]
        data to be infilled
    modified_path : pathlib.Path
        path to user modified data
    uneditable_columns : Optional[dict[str, list[str]]], optional
        columns that the user is forbidden to edit, by default None

    Returns
    -------
    dict[str, pd.DataFrame]
        data with user fixes integrated

    Raises
    ------
    FileNotFoundError
        if inputted file path does not exist
    ValueError
        if columns listed in uneditable have been modified 
    """
    # Read in the Excel file containing the subset of the data
    infilled_data = {}
    if not os.path.exists(modified_path):
        raise FileNotFoundError(f"{modified_path} does not exisit")

    for key, value in data.items():

        data_subset = parser.parse_sheet(modified_path, key, 0)

        # removes "" from luc lists
        data_subset["existing_land_use"] = data_subset[
            "existing_land_use"].apply(lambda x: [
                item for item in x if item != ""] if isinstance(x, list) else x)
        data_subset["proposed_land_use"] = data_subset[
            "proposed_land_use"].apply(lambda x: [
                item for item in x if item != ""] if isinstance(x, list) else x)

        if uneditable_columns is None:
            infilled_data[key] = value.copy()
            # Update the values in the original dataframe with the values in the subset
            infilled_data[key].update(data_subset)
            continue

        # Check whether any values in the specified columns have been changed
        else:
            if not value[uneditable_columns[key]].equals(data_subset[uneditable_columns[key]]):
                infilled_data[key] = value.copy()
                # Update the values in the original dataframe with the values in the subset
                infilled_data[key].update(data_subset)
            else:
                raise ValueError(
                    f"values in {uneditable_columns[key]} within {modified_path} have been modified, these values must remain constant")
    return infilled_data

def implement_user_fixes(
    config: inputs.DLitConfig,
    dlog_data: global_classes.DLogData,
    auxiliary_data: global_classes.AuxiliaryData,
    plot_graphs: bool,
    )->Optional[global_classes.DLogData]:
    """intergrates user fixes into data

    handles user preferences writes the file for the user to edit,
    reads in user fixes and outputs the data with the user fixes
    integrated. outputs None if user wishes to end program 

    Parameters
    ----------
    config : inputs.DLitConfig
        config object from dlit config file
    dlog_data : global_classes.DLogData
        data to infill
    auxiliary_data : global_classes.AuxiliaryData
        auxiliary data from parser

    Returns
    -------
    Optional[global_classes.DLogData]
        infilled data, None if user wishes to end program
    """    

    #determines if user wishes to infill using exisiting file
    if os.path.exists(config.user_input_path):
        modification_file_ready = utilities.y_n_user_input(
            f"A file already exists at {config.user_input_path}."
            " Does this contain the fixes you wish to implement? (Y/N)\n")
    else:
        modification_file_ready = False
    
    #adds filter columns without producing report

    if modification_file_ready:
        user_changes = True
        LOG.info(f"Existing file {config.user_input_path} set as user infill input.")
    else:
        report_path = config.output_folder / "initial_data_quality_report.xlsx"
        analyse.data_report(
            dlog_data,
            report_path,
            config.output_folder,
            auxiliary_data,
            plot_graphs,   
            True,
        )

        LOG.info(f"Intial data quality report saved as {report_path}")

        #checks if user wishes to infill data
        user_changes = utilities.y_n_user_input("Do you wish to "
            "manually fix data before it is infilled? (Y/N)\n")

        if user_changes:
            if os.path.exists(config.user_input_path):
                LOG.info("Creating file for user to edit.")
                #pauses to allow user to save existing file
                input(f"Overwriting {config.user_input_path}, if you wish to store any changes made"
                    ", please make a copy with a different name and press enter, otherwise press"
                    " enter.")

            user_input_file_builder(
                config.user_input_path, dlog_data)

            #allows user to end program to to edit data
            end_program = utilities.y_n_user_input(f"A file has been created at "
                    f"{config.user_input_path} for you to manually infill data. Would "
                    "you like to end the program and rerun when you have finished? Y "
                    "(end the program, modify the data then rerun) or N (data has been"
                    " modified)\n")

            if end_program:
                LOG.info("Ending program")
                return None

    if user_changes:
        LOG.info("Implementing user fixes")
        infilled_data = infill_user_inputs(
            dict((k, utilities.to_dict(dlog_data)[k]) for k in (["residential","employment", "mixed"])),
            config.user_input_path)
        infilled_data = utilities.to_dlog_data(infilled_data, dlog_data.lookup)
    return infilled_data

@utilities.output_file_checks
def create_user_changes_audit(
    file_path:pathlib.Path,
    input_modified: global_classes.DLogData,
    input_original:global_classes.DLogData,
    )->None:
    modified = utilities.to_dict(input_modified)
    original = utilities.to_dict(input_original)
    workbook = openpyxl.Workbook()

    # Iterate over the dictionaries
    for key, modified_df_ in modified.items():

        LOG.info(f"Writing {key} sheet")

        # convert lists to strings, otherwise openpyxl throws a wobbly
        original_df = original[key].applymap(convert_list_to_string)
        modified_df = modified_df_.applymap(convert_list_to_string)
        
        # Create a new sheet
        sheet = workbook.create_sheet(key)
        sheet.append(modified_df.columns.tolist())
        sheet_index = 2

        # Iterate over the rows of modified_df
        for index, modified_row in modified_df.iterrows():
            original_row = original_df.loc[index]
            # Compare the values in each column
            altered = False
            altered_columns = []
            for column in modified_df.columns:
                if pd.isna(original_df[column][index]) and pd.isna(modified_df[column][index]):
                    continue

                # If the column is not a list, compare the values directly
                elif original_df[column][index] != modified_df[column][index]:
                    altered = True
                    altered_columns.append(column)
                    break

            # If the row has been modified, add it to the sheet and color the cells red


            if altered:
                for i, value in enumerate(modified_row):
                    cell = sheet.cell(row=sheet_index, column=i+1)
                    if pd.isna(value):
                        cell.value = None
                    else:
                        cell.value = value
                    if modified_df.columns[i] in altered_columns:
                        cell.font = openpyxl.styles.Font(color='FF0000')
                sheet_index += 1

    # Save the workbook
    workbook.save(file_path)

def convert_list_to_string(value):
    if isinstance(value, list):
        return ', '.join(value)
    return value